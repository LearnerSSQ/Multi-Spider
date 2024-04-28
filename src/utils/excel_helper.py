import os
import sys
from typing import Tuple

from openpyxl.reader.excel import load_workbook
from rich.console import Console
from rich.table import Table

from .input_validator import InputValidator


class ExcelHelper:
    def __init__(self):
        self.console = Console()
        self.input_validator = InputValidator()

    def select_excel_file(self) -> str:
        excel_files = [file for file in os.listdir('.') if file.endswith('.xlsx')]

        if not excel_files:
            self.console.print("[bold red]当前目录下没有找到任何 .xlsx 文件[/bold red]")
            input('按任意键退出...')
            sys.exit(1)

        table = Table(title="Excel 文件列表", show_header=True, header_style="green")
        table.add_column("序号", justify="center", style="yellow", no_wrap=True)
        table.add_column("文件名", justify="left", style="yellow")

        for idx, excel_file in enumerate(excel_files, start=1):
            table.add_row(str(idx), excel_file)

        self.console.print(table)
        select_idx = self.input_validator.get_valid_input("请输入要读取的 Excel 文件序号: ", 1, len(excel_files))

        return excel_files[select_idx - 1]

    def select_excel_columns(self, file_path: str) -> Tuple[int, int]:
        wb = load_workbook(filename=file_path)
        sheet = wb.active

        table = Table(title="Excel 文件标题行", show_header=True, header_style="green")
        table.add_column("列编号", justify="center", style="yellow", no_wrap=True)
        table.add_column("列标题", justify="left", style="yellow")

        for idx, cell in enumerate(sheet[1], start=1):
            table.add_row(str(idx), str(cell.value))

        self.console.print(table)

        read_col_num = self.input_validator.get_valid_input("请输入需要读取数据的列的起始位置: ", 1, sheet.max_column)
        # self.console.print('覆写的内容为N行3列,内容为"新增用户名, 新增联系号码1, 新增联系号码2",只需要选取最左侧的列即可\n举例: 如果选取了第4列，那么数据将会覆盖到第4, 5, 6列')
        write_col_num = self.input_validator.get_valid_input("请输入需要进行覆写的列的起始位置: ", 1, sheet.max_column,
                                                             exclude=read_col_num)

        return read_col_num, write_col_num

    @staticmethod
    def read_excel_columns(read_excel_file: str, start_col_num: int, end_col_num: int) -> list:
        result = []

        wb = load_workbook(read_excel_file)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, min_col=start_col_num, max_col=end_col_num, values_only=True):
            if row[0] is not None:
                result.append(row)
            else:
                break

        return result

    @staticmethod
    def write_to_excel(selected_excel_file: str, start_row: int, start_column: int, data_to_write: list) -> None:
        wb = load_workbook(filename=selected_excel_file)
        sheet = wb.active

        for row_index, row_data in enumerate(data_to_write, start=start_row):
            for column_index, cell_value in enumerate(row_data, start=start_column):
                sheet.cell(row=row_index, column=column_index, value=cell_value)

        wb.save(selected_excel_file)

    def check_and_close_excel_file(self, file_path: str) -> None:
        try:
            wb = load_workbook(filename=file_path)
            wb.close()
            self.console.print(f"[green]文件 {file_path} 未被打开,或已成功关闭,\n可以进行下一步操作[/green]")

        except PermissionError:
            self.console.print(f"[bold red]文件 {file_path} 当前被打开,请关闭它后再试[/bold red]")
            input('按任意键退出...')
            sys.exit(1)

        except Exception as e:
            self.console.print(f"[bold red]检查文件时出现未知错误: {e}[/bold red]")
            input('按任意键退出...')
            sys.exit(1)
